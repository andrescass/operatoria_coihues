import openpyxl
from openpyxl.styles import Fill, PatternFill, Font
from openpyxl.styles.colors import Color
import sys
from tkinter import *
from tkinter import filedialog, messagebox

bg_color_main = "FFE599"

class Vendedor:
    def __init__(self, name):
        self.name = name
        self.products = []
        self.prices = {} # <product, precio>
        self.vendidos = {} # <product, cantidad>
        self.total = 0
        

class Comprador:
    def __init__(self, name):
        self.name = name
        self.stamps = []
        self.products = [] # <(stamp, product, cantidad)>
        self.total = 0
        self.bono = 0

def process():
    inFileName = fileNameEntry.get()
    if inFileName == '':
        messagebox.showinfo(title="Error", message="Debe seleccionar un archivo")
        return
    
    if inFileName.split('.')[-1] != 'xlsx':
        messagebox.showinfo(title="Error", message="Debe seleccionar un archivo xlsx")
        return
    inFile = openpyxl.load_workbook(inFileName) #abrir documento de entrada
    wSheet = inFile.get_sheet_by_name('Respuestas de formulario 1') #trabajo sobre la primer hoja

    #armar vector de vendedores
    print("Leyendo productores")
    vendedores = {}
    col_i = 0
    mensaje_colaboracion = "Registrá, por favor, tu   NOMBRE   Y   APELLIDO  (registrá el que vas a usar habitualmente con nosotros)"
    mensaje_limite = "Registrá"
    for c in wSheet.iter_cols(min_col=2, values_only=True):
        #if c[0] != mensaje_colaboracion:
        #if c[0].split(',')[0] != mensaje_limite:
        if mensaje_limite not in c[0]:
            if c[0] and ("\"1\"" not in c[0]) and ("$" in c[0] or "=" in c[0]):
                if '.' in c[0].split('[')[0]:
                    nombre = c[0].split('.')[0]
                else:
                    nombre = c[0].split('[')[0]
                if nombre not in vendedores:
                    v = Vendedor(nombre)
                    v.products.append(c[0])
                    if '$' in c[0]:
                        v.prices[c[0]] = int(c[0].split('$')[1].split(']')[0])
                    elif '=' in c[0]:
                        v.prices[c[0]] = int(c[0].split('=')[1].split(']')[0])
                    else: 
                        v.prices[c[0]] = -1
                    v.vendidos[c[0]] = 0
                    vendedores[nombre] = v
                else:
                    vendedores[nombre].products.append(c[0])
                    if '$' in c[0]:
                        vendedores[nombre].prices[c[0]] = int(c[0].split('$')[1].split(']')[0])
                    elif '=' in c[0]:
                        vendedores[nombre].prices[c[0]] = int(c[0].split('=')[1].split(']')[0])
                    else: 
                        vendedores[nombre].prices[c[0]] = -1
                    vendedores[nombre].vendidos[c[0]] = 0
        else:
            break
    
    #armar vector de compradores)
    print("Leyendo compradores")
    mensaje_nombre = "Registrá, por favor, tu   NOMBRE   Y   APELLIDO  (registrá el que vas a usar habitualmente con nosotros)"
    celda_nombre = 0
    for c in wSheet.iter_cols(min_col=1, values_only = True):
        if c[0].split(',')[0] != mensaje_limite:
            celda_nombre += 1
        else:
            break
    
    # Buscar columna de bono
    celda_bono = 0
    for c in wSheet.iter_cols(min_col=1, values_only = True):
        if "BONO" not in c[0]:
            celda_bono += 1
        else:
            break

    compradores = {}
    row_counter = 0
    for r in wSheet.iter_rows(min_row=2, values_only=True):
        #print(r)
        row_counter += 1
        nombre = r[celda_nombre]
        if nombre not in compradores:
            comprador = Comprador(nombre)
            comprador.stamps.append(r[0])
            for c in range(1, len(r)):
                #if wSheet.cell(row=1, column=c).value == mensaje_colaboracion:
                if wSheet.cell(row=1, column=c+1).value and ("\"1\"" not in wSheet.cell(row=1, column=c+1).value) and ("$" in wSheet.cell(row=1, column=c+1).value or "=" in wSheet.cell(row=1, column=c+1).value):
                    if 'Registrá' in wSheet.cell(row=1, column=c+1).value:
                        break
                    elif r[c] != None:
                        try:
                            stamp = r[0].strftime("%d-%m-%y")
                            prod = wSheet.cell(row=1, column=c+1).value
                            nombre_vendedor = prod.split('.')[0] if '.' in prod.split('[')[0] else prod.split('[')[0]
                            comprador.products.append((stamp, prod, int(r[c])))
                            try:
                                vendedores[nombre_vendedor].vendidos[prod] += int(r[c])
                            except KeyError:
                                print("Error en {0} {1}".format(nombre_vendedor, c))
                        except ValueError:
                            print("Caracter no numérico en fila {0}".format(wSheet.cell(row=row_counter, column=c+1).row + 1))

            compradores[nombre] = comprador
        else:
            compradores[nombre].stamps.append(r[0])
            for c in range(1, len(r)):
                #if wSheet.cell(row=1, column=c).value == mensaje_colaboracion:
                if mensaje_limite in wSheet.cell(row=1, column=c+1).value:
                    break
                elif r[c] != None:
                    stamp = r[0].strftime("%d-%m-%y")
                    prod = wSheet.cell(row=1, column=c+1).value
                    nombre_vendedor = prod.split('.')[0] if '.' in prod.split('[')[0] else prod.split('[')[0]
                    compradores[nombre].products.append((stamp, prod, int(r[c])))
                    vendedores[nombre_vendedor].vendidos[prod] += int(r[c])

        if celda_bono > 0 and r[celda_bono] != None:
            if not isinstance(r[celda_bono], int) and not isinstance(r[celda_bono], float):
                if '-' in r[celda_bono]:
                    if '$' in r[celda_bono].split('-')[1]:
                        compradores[nombre].bono -= int(r[celda_bono].split('-')[1].split('$')[1])
                    else:
                        compradores[nombre].bono -= int(r[celda_bono].split('-')[1])
                elif '$' in r[celda_bono]:
                    if '-' in r[celda_bono].split('$')[1]:
                        compradores[nombre].bono -= int(r[celda_bono].split('$')[1].split('-')[1])
                    else:
                        compradores[nombre].bono -= int(r[celda_bono].split('$')[1])
            else:
                compradores[nombre].bono -= int(r[celda_bono])
            print("{0}, bono {1}".format(compradores[nombre].name, compradores[nombre].bono))


    #Imprimir vendedores
    print("Produciendo salida")
    print("Imprimiendo productores")
    sheet_vendedores = "resumen productores"
    if sheet_vendedores not in inFile.get_sheet_names():
        inFile.create_sheet(sheet_vendedores)
        vend_sheet = inFile.get_sheet_by_name(sheet_vendedores)
        row_idx = 1
        vend_sheet.cell(row=row_idx, column=1).value = "Nombre del Productor"
        vend_sheet.cell(row=row_idx, column=2).value = "Precios"
        vend_sheet.cell(row=row_idx, column=3).value = "Cantidades"
        row_idx += 1
        v_proc = 0

        for v in sorted(vendedores.keys()):
            totales = 0
            for p in vendedores[v].products:
                vend_sheet.cell(row=row_idx, column=1).value = p
                vend_sheet.cell(row=row_idx, column=2).value = "{:.2f}".format(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                vend_sheet.cell(row=row_idx, column=3).value = vendedores[v].vendidos[p]
                totales += vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p]
                row_idx += 1
            vend_sheet.cell(row=row_idx, column=2).value = "Suma"
            vend_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=bg_color_main)
            vend_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
            vend_sheet.cell(row=row_idx, column=3).value = "{:.2f}".format(totales)
            vend_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
            vend_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
            row_idx += 1
            v_proc += 1
            vendedores[v].total = totales
            print("Procesado {0} %                    \r".format(int(v_proc/len(vendedores)*100)), end="\r")
        inFile.save(filename=inFileName)
    else:
        #user_option = input("\n \n Ya existe la hoja Resumen Productores, desea sobreescribirla? [Si/No]")
        warningMessage = "Ya existe la hoja Resumen Productores, desea sobreescribirla?"
        user_option = messagebox.askyesno(title = "Hoja existente", message=warningMessage)
        #if user_option == "Si" or user_option == "si":
        if user_option:
            #inFile.create_sheet(sheet_vendedores)
            vend_sheet = inFile.get_sheet_by_name(sheet_vendedores)
            inFile.remove_sheet(vend_sheet)
            inFile.create_sheet(sheet_vendedores)
            vend_sheet = inFile.get_sheet_by_name(sheet_vendedores)

            row_idx = 1
            vend_sheet.cell(row=row_idx, column=1).value = "Nombre del Productor"
            vend_sheet.cell(row=row_idx, column=2).value = "Precios"
            vend_sheet.cell(row=row_idx, column=3).value = "Cantidades"
            row_idx += 1
            v_proc = 0

            for v in sorted(vendedores.keys()):
                totales = 0
                for p in vendedores[v].products:
                    vend_sheet.cell(row=row_idx, column=1).value = p
                    vend_sheet.cell(row=row_idx, column=2).value = "{:.2f}".format(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                    vend_sheet.cell(row=row_idx, column=3).value = vendedores[v].vendidos[p]
                    totales += vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p]
                    row_idx += 1
                vend_sheet.cell(row=row_idx, column=2).value = "Suma"
                vend_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=bg_color_main)
                vend_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                vend_sheet.cell(row=row_idx, column=3).value = "{:.2f}".format(totales)
                vend_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
                vend_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                row_idx += 1
                v_proc += 1
                vendedores[v].total = totales
                print("Procesado {0} %                    \r".format(int(v_proc/len(vendedores)*100)), end="\r")
            inFile.save(filename=inFileName)
        else:
            print("Continuando sin sobreescribir")

    #Imprimir compradores
    sheet_compradores = "resumen compradores"
    print("Imprimiendo compradores")
    if sheet_compradores not in inFile.get_sheet_names():
        inFile.create_sheet(sheet_compradores)
        comp_sheet = inFile.get_sheet_by_name(sheet_compradores)
        row_idx = 1
        c_proc = 0
        
        for c in compradores.values():
            comp_sheet.cell(row=row_idx, column=1).value = c.name
            comp_sheet.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=bg_color_main)
            comp_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
            #comp_sheet.cell(row=row_idx, column=2).value = "Cantidad"
            comp_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=bg_color_main)
            #comp_sheet.cell(row=row_idx, column=3).value = "Precio"
            comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
            row_idx += 1
            
            total_c = 0
            for p in sorted(c.products, key=lambda x: x[1]):
                comp_sheet.cell(row=row_idx, column=1).value = p[1]
                comp_sheet.cell(row=row_idx, column=2).value = p[2]
                nombre_v = p[1].split('.')[0] if '.' in p[1].split('[')[0] else p[1].split('[')[0]
                p_precio = vendedores[nombre_v].prices[p[1]]
                comp_sheet.cell(row=row_idx, column=3).value = int(p[2]*p_precio)
                row_idx += 1
                total_c += int(p[2]*p_precio)
            total_c += c.bono

            if c.bono != 0:
                comp_sheet.cell(row=row_idx, column=2).value = "Bono"
                comp_sheet.cell(row=row_idx, column=3).value = c.bono
                row_idx += 1
            comp_sheet.cell(row=row_idx, column=2).value = "Suma"
            comp_sheet.cell(row=row_idx, column=3).value = total_c
            row_idx += 1
            comp_sheet.cell(row=row_idx, column=2).value = "Redondeo"
            comp_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
            comp_sheet.cell(row=row_idx, column=3).value = total_c - (total_c%10)
            comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
            comp_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
            row_idx += 1
            c_proc += 1
            c.total = total_c - (total_c%10)
            print("Procesado {0}%                    \r".format(int(c_proc/len(compradores)*100)), end="\r")
        inFile.save(filename=inFileName)
    else:
        #user_option = input("\n \n Ya existe la hoja Resumen Compradores, desea sobreescribirla? [Si/No]")
        #if user_option == "Si" or user_option == "si":
        warningMessage = "Ya existe la hoja Resumen Compradores, desea sobreescribirla?"
        user_option = messagebox.askyesno(title = "Hoja existente", message=warningMessage)
        if user_option:
            comp_sheet = inFile.get_sheet_by_name(sheet_compradores)
            inFile.remove_sheet(comp_sheet)
            inFile.create_sheet(sheet_compradores)
            comp_sheet = inFile.get_sheet_by_name(sheet_compradores)

            row_idx = 1
            c_proc = 0
            
            for c in compradores.values():
                comp_sheet.cell(row=row_idx, column=1).value = c.name
                comp_sheet.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=bg_color_main)
                comp_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
                #comp_sheet.cell(row=row_idx, column=2).value = "Cantidad"
                comp_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=bg_color_main)
                #comp_sheet.cell(row=row_idx, column=3).value = "Precio"
                comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
                row_idx += 1
                
                total_c = 0
                for p in sorted(c.products, key=lambda x: x[1]):
                    comp_sheet.cell(row=row_idx, column=1).value = p[1]
                    comp_sheet.cell(row=row_idx, column=2).value = p[2]
                    nombre_v = p[1].split('.')[0] if '.' in p[1].split('[')[0] else p[1].split('[')[0]
                    p_precio = vendedores[nombre_v].prices[p[1]]
                    comp_sheet.cell(row=row_idx, column=3).value = int(p[2]*p_precio)
                    row_idx += 1
                    total_c += int(p[2]*p_precio)
                
                total_c += c.bono

                if c.bono != 0:
                    comp_sheet.cell(row=row_idx, column=2).value = "Bono"
                    comp_sheet.cell(row=row_idx, column=3).value = c.bono
                    row_idx += 1                
                comp_sheet.cell(row=row_idx, column=2).value = "Suma"
                comp_sheet.cell(row=row_idx, column=3).value = total_c
                row_idx += 1
                comp_sheet.cell(row=row_idx, column=2).value = "Redondeo"
                comp_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                comp_sheet.cell(row=row_idx, column=3).value = total_c - (total_c%10)
                comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=bg_color_main)
                comp_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                row_idx += 1
                c_proc += 1
                c.total = total_c - (total_c%10)
            print("Procesado {0}%                    \r".format(int(c_proc/len(compradores)*100)), end="\r")
            inFile.save(filename=inFileName)
        else:
            print("Saliendo sin sobreescribir")
    
    if(isFinalProcess): #cierre de operatoria
        print("Realizando resumen final de compradores")
        sheet_compradores_final = "Resumen final compradores"
        if sheet_compradores_final not in inFile.get_sheet_names():
            inFile.create_sheet(sheet_compradores_final)
        else:
            comp_sheet_final = inFile.get_sheet_by_name(sheet_compradores_final)
            inFile.remove_sheet(comp_sheet_final)
            inFile.create_sheet(sheet_compradores_final)

        comp_sheet_final = inFile.get_sheet_by_name(sheet_compradores_final)
        row_idx = 1

        for c in compradores.values():
            comp_sheet_final.cell(row=row_idx, column=1).value = c.name
            comp_sheet_final.cell(row=row_idx, column=2).value = "{:.2f}".format(c.total)
            row_idx+=1

        print("Realizando resumen final de productores")
        sheet_vendedores_final = "Resumen final productores"
        if sheet_vendedores_final not in inFile.get_sheet_names():
            inFile.create_sheet(sheet_vendedores_final)
        else:
            vend_sheet_final = inFile.get_sheet_by_name(sheet_vendedores_final)
            inFile.remove_sheet(vend_sheet_final)
            inFile.create_sheet(sheet_vendedores_final)
        
        vend_sheet_final = inFile.get_sheet_by_name(sheet_vendedores_final)

        row_idx = 1

        for v in sorted(vendedores.keys()):
            vend_sheet_final.cell(row=row_idx, column=1).value = vendedores[v].name
            vend_sheet_final.cell(row=row_idx, column=2).value = "{:.2f}".format(vendedores[v].total)
            row_idx+=1

        inFile.save(filename=inFileName)
    
    messagebox.showinfo(title = "Procesamiento terminado", message="Procesamiento finalizado")

def openFileWindow():
    path = filedialog.askopenfilename(initialdir="./", title="Select file",
                    filetypes=(("excel", "*.xlsx"),("all files", "*.*")))
    fileNameEntry.delete(0, "end")
    fileNameEntry.insert(0, path)

if __name__ == "__main__":
    # Armo la ventana
    root = Tk()
    root.geometry("360x150")
    root.title("Operatoria")
    mainFrame = Frame(root)
    mainFrame.grid(column=0, row=0, sticky=(N, W, E, S))

    leftFrame = Frame(root)
    #leftFrame.pack(side = LEFT)
    rightFrame = Frame(root)
    #rightFrame.pack(side = RIGHT)

    label = Label(mainFrame, text = "Seleccione el archivo")
    label.grid(column = 1, row = 0)
    #label.pack()

    fileNameEntry = Entry(mainFrame, width=50)
    fileNameEntry.grid(column = 0, columnspan = 2, row = 1)
    #fileNameEntry.pack(padx=5)
    openFileBut = Button(mainFrame, text = "Abrir", command = openFileWindow)
    openFileBut.grid(column = 2, row = 1)
    #openFileBut.pack(padx=5)

    isFinalProcess = BooleanVar()
    finalCheck = Checkbutton(mainFrame, text="Final de operatoria", variable=isFinalProcess)
    finalCheck.grid(column = 0, row = 2)

    processBut = Button(mainFrame, text = "Procesar", command = process)
    processBut.grid(column = 1, row = 2)

    
    #processBut.pack(padx=20, pady = 5)

    root.mainloop()


        
