import openpyxl
from openpyxl.styles import Fill, PatternFill, Font
from openpyxl.styles.colors import Color
import sys

class Vendedor:
    def __init__(self, name):
        self.name = name
        self.products = []
        self.prices = {} # <product, precio>
        self.vendidos = {} # <product, cantidad>

class Comprador:
    def __init__(self, name):
        self.name = name
        self.stamps = []
        self.products = [] # <(stamp, product, cantidad)>


if __name__ == "__main__":
    if len(sys.argv) > 1:
        inFileName = sys.argv[1]
        inFile = openpyxl.load_workbook(inFileName) #abrir documento de entrada
        wSheet = inFile.get_sheet_by_name('Respuestas de formulario 1') #trabajo sobre la primer hoja

        #armar vector de vendedores
        print("Leyendo productores")
        vendedores = {}
        col_i = 0
        mensaje_colaboracion = "Registrá, por favor, tu   NOMBRE   Y   APELLIDO  (registrá el que vas a usar habitualmente con nosotros)"
        for c in wSheet.iter_cols(min_col=2, values_only=True):
            if c[0] != mensaje_colaboracion:
                nombre = c[0].split('[')[0]
                if nombre not in vendedores:
                    v = Vendedor(nombre)
                    v.products.append(c[0])
                    if '$' in c[0]:
                        v.prices[c[0]] = int(c[0].split('$')[1].split(']')[0])
                    elif '=' in c[0]:
                        v.prices[c[0]] = int(c[0].split('=')[1].split(']')[0])
                    else: 
                        v.prices[c[0]] = -1
                    v.vendidos[c[0]] = 0
                    vendedores[nombre] = v
                else:
                    vendedores[nombre].products.append(c[0])
                    if '$' in c[0]:
                        vendedores[nombre].prices[c[0]] = int(c[0].split('$')[1].split(']')[0])
                    elif '=' in c[0]:
                        vendedores[nombre].prices[c[0]] = int(c[0].split('=')[1].split(']')[0])
                    else: 
                        vendedores[nombre].prices[c[0]] = -1
                    vendedores[nombre].vendidos[c[0]] = 0
            else:
                break
        
        #armar vector de compradores)
        print("Leyendo compradores")
        mensaje_nombre = "Registrá, por favor, tu   NOMBRE   Y   APELLIDO  (registrá el que vas a usar habitualmente con nosotros)"
        celda_nombre = 0
        for c in wSheet.iter_cols(min_col=1, values_only = True):
            if c[0] != mensaje_nombre:
                celda_nombre += 1
            else:
                break

        compradores = {}
        row_counter = 0
        for r in wSheet.iter_rows(min_row=2, values_only=True):
            #print(r)
            row_counter += 1
            nombre = r[celda_nombre]
            if nombre not in compradores:
                comprador = Comprador(nombre)
                comprador.stamps.append(r[0])
                for c in range(1, len(r)):
                    #if wSheet.cell(row=1, column=c).value == mensaje_colaboracion:
                    if 'NOMBRE   Y   APELLIDO' in wSheet.cell(row=1, column=c+1).value:
                        break
                    elif r[c] != None:
                        try:
                            stamp = r[0].strftime("%d-%m-%y")
                            prod = wSheet.cell(row=1, column=c+1).value
                            nombre_vendedor = prod.split('[')[0]
                            comprador.products.append((stamp, prod, int(r[c])))
                            try:
                                vendedores[nombre_vendedor].vendidos[prod] += int(r[c])
                            except KeyError:
                                print("Error en {0} {1}".format(nombre_vendedor, c))
                        except ValueError:
                            print("Caracter no numérico en fila {0}".format(wSheet.cell(row=row_counter, column=c+1).row + 1))


                compradores[nombre] = comprador
            else:
                compradores[nombre].stamps.append(r[0])
                for c in range(1, len(r)):
                    #if wSheet.cell(row=1, column=c).value == mensaje_colaboracion:
                    if 'NOMBRE   Y   APELLIDO' in wSheet.cell(row=1, column=c+1).value:
                        break
                    elif r[c] != None:
                        stamp = r[0].strftime("%d-%m-%y")
                        prod = wSheet.cell(row=1, column=c+1).value
                        nombre_vendedor = prod.split('[')[0]
                        compradores[nombre].products.append((stamp, prod, int(r[c])))
                        vendedores[nombre_vendedor].vendidos[prod] += int(r[c])

        #Imprimir vendedores
        print("Produciendo salida")
        print("Imprimiendo productores")
        sheet_vendedores = "resumen productores"
        if sheet_vendedores not in inFile.get_sheet_names():
            inFile.create_sheet(sheet_vendedores)
            vend_sheet = inFile.get_sheet_by_name(sheet_vendedores)
            row_idx = 1
            vend_sheet.cell(row=row_idx, column=1).value = "Nombre del Productor"
            vend_sheet.cell(row=row_idx, column=2).value = "Precios"
            vend_sheet.cell(row=row_idx, column=3).value = "Cantidades"
            row_idx += 1
            v_proc = 0

            for v in sorted(vendedores.keys()):
                totales = 0
                for p in vendedores[v].products:
                    vend_sheet.cell(row=row_idx, column=1).value = p
                    vend_sheet.cell(row=row_idx, column=2).value = int(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                    vend_sheet.cell(row=row_idx, column=3).value = vendedores[v].vendidos[p]
                    totales += int(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                    row_idx += 1
                vend_sheet.cell(row=row_idx, column=2).value = "Suma"
                vend_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor="FFFF00")
                vend_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                vend_sheet.cell(row=row_idx, column=3).value = totales
                vend_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                vend_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                row_idx += 1
                v_proc += 1
                print("Procesado {0} %                    \r".format(int(v_proc/len(vendedores)*100)), end="\r")
            inFile.save(filename=inFileName)
        else:
            user_option = input("\n \n Ya existe la hoja Resumen Productores, desea sobreescribirla? [Si/No]")
            if user_option == "Si" or user_option == "si":
                #inFile.create_sheet(sheet_vendedores)
                vend_sheet = inFile.get_sheet_by_name(sheet_vendedores)
                row_idx = 1
                vend_sheet.cell(row=row_idx, column=1).value = "Nombre del Productor"
                vend_sheet.cell(row=row_idx, column=2).value = "Precios"
                vend_sheet.cell(row=row_idx, column=3).value = "Cantidades"
                row_idx += 1
                v_proc = 0

                for v in sorted(vendedores.keys()):
                    totales = 0
                    for p in vendedores[v].products:
                        vend_sheet.cell(row=row_idx, column=1).value = p
                        vend_sheet.cell(row=row_idx, column=2).value = int(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                        vend_sheet.cell(row=row_idx, column=3).value = vendedores[v].vendidos[p]
                        totales += int(vendedores[v].prices[p]/1.05 * vendedores[v].vendidos[p])
                        row_idx += 1
                    vend_sheet.cell(row=row_idx, column=2).value = "Suma"
                    vend_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor="FFFF00")
                    vend_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                    vend_sheet.cell(row=row_idx, column=3).value = totales
                    vend_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                    vend_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                    row_idx += 1
                    v_proc += 1
                    print("Procesado {0} %                    \r".format(int(v_proc/len(vendedores)*100)), end="\r")
                inFile.save(filename=inFileName)
            else:
                print("Continuando sin sobreescribir")

        #Imprimir compradores
        sheet_compradores = "resumen compradores"
        print("Imprimiendo compradores")
        if sheet_compradores not in inFile.get_sheet_names():
            inFile.create_sheet(sheet_compradores)
            comp_sheet = inFile.get_sheet_by_name(sheet_compradores)
            row_idx = 1
            c_proc = 0
            
            for c in compradores.values():
                comp_sheet.cell(row=row_idx, column=1).value = c.name
                comp_sheet.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor="FFFF00")
                comp_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
                comp_sheet.cell(row=row_idx, column=2).value = "Cantidad"
                comp_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor="FFFF00")
                comp_sheet.cell(row=row_idx, column=3).value = "Precio"
                comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                row_idx += 1
                
                total_c = 0
                for p in c.products:
                    comp_sheet.cell(row=row_idx, column=1).value = p[1]
                    comp_sheet.cell(row=row_idx, column=2).value = p[2]
                    nombre_v = p[1].split('[')[0]
                    p_precio = vendedores[nombre_v].prices[p[1]]
                    comp_sheet.cell(row=row_idx, column=3).value = int(p[2]*p_precio)
                    row_idx += 1
                    total_c += int(p[2]*p_precio)
                
                comp_sheet.cell(row=row_idx, column=2).value = "Suma"
                comp_sheet.cell(row=row_idx, column=3).value = total_c
                row_idx += 1
                comp_sheet.cell(row=row_idx, column=2).value = "Redondeo"
                comp_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                comp_sheet.cell(row=row_idx, column=3).value = total_c - (total_c%10)
                comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                comp_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                row_idx += 1
                c_proc += 1
                print("Procesado {0}%                    \r".format(int(c_proc/len(compradores)*100)), end="\r")
            inFile.save(filename=inFileName)
        else:
            user_option = input("\n \n Ya existe la hoja Resumen Compradores, desea sobreescribirla? [Si/No]")
            if user_option == "Si" or user_option == "si":
                comp_sheet = inFile.get_sheet_by_name(sheet_compradores)
                row_idx = 1
                c_proc = 0
                
                for c in compradores.values():
                    comp_sheet.cell(row=row_idx, column=1).value = c.name
                    comp_sheet.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor="FFFF00")
                    comp_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
                    comp_sheet.cell(row=row_idx, column=2).value = "Cantidad"
                    comp_sheet.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor="FFFF00")
                    comp_sheet.cell(row=row_idx, column=3).value = "Precio"
                    comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                    row_idx += 1
                    
                    total_c = 0
                    for p in c.products:
                        comp_sheet.cell(row=row_idx, column=1).value = p[1]
                        comp_sheet.cell(row=row_idx, column=2).value = p[2]
                        nombre_v = p[1].split('[')[0]
                        p_precio = vendedores[nombre_v].prices[p[1]]
                        comp_sheet.cell(row=row_idx, column=3).value = int(p[2]*p_precio)
                        row_idx += 1
                        total_c += int(p[2]*p_precio)
                    
                    comp_sheet.cell(row=row_idx, column=2).value = "Suma"
                    comp_sheet.cell(row=row_idx, column=3).value = total_c
                    row_idx += 1
                    comp_sheet.cell(row=row_idx, column=2).value = "Redondeo"
                    comp_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
                    comp_sheet.cell(row=row_idx, column=3).value = total_c - (total_c%10)
                    comp_sheet.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFFF00")
                    comp_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
                    row_idx += 1
                    c_proc += 1
                print("Procesado {0}%                    \r".format(int(c_proc/len(compradores)*100)), end="\r")
                inFile.save(filename=inFileName)
            else:
                print("Saliendo sin sobreescribir")
      
    else:
        print("Indicar el nombre del archivo de entrada")
